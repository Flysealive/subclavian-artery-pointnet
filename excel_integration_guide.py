#!/usr/bin/env python3

import pandas as pd
import numpy as np
import os

def prepare_excel_template():
    """
    Create an Excel template with proper formatting for easy data entry
    """
    # Read existing labels
    df = pd.read_csv('classification_labels.csv')
    
    # Extract model ID for easier matching
    df['Model_ID'] = df['filename'].str.extract(r'3DModel(\d+)')
    
    # Reorder columns for clarity
    df = df[['Model_ID', 'filename', 'label']]
    
    # Add measurement columns with descriptions
    df['left_subclavian_diameter_mm'] = np.nan
    df['aortic_arch_diameter_mm'] = np.nan
    df['angle_degrees'] = np.nan
    df['notes'] = ''  # Optional notes column
    
    # Create Excel writer with formatting
    output_file = 'measurement_template.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Measurements', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Measurements']
        
        # Add header formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, 8):  # Columns A-G
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15  # Model_ID
        worksheet.column_dimensions['B'].width = 25  # filename
        worksheet.column_dimensions['C'].width = 10  # label
        worksheet.column_dimensions['D'].width = 25  # left_subclavian_diameter_mm
        worksheet.column_dimensions['E'].width = 25  # aortic_arch_diameter_mm
        worksheet.column_dimensions['F'].width = 15  # angle_degrees
        worksheet.column_dimensions['G'].width = 30  # notes
        
        # Add instructions sheet
        instructions_df = pd.DataFrame({
            'Instructions': [
                'MEASUREMENT GUIDELINES',
                '',
                '1. MEASUREMENTS REQUIRED:',
                '   - Left Subclavian Diameter: Measure in millimeters (typical range: 5-15 mm)',
                '   - Aortic Arch Diameter: Measure in millimeters (typical range: 20-40 mm)',
                '   - Angle: Angle between left subclavian artery and aortic arch in degrees (typical range: 30-150°)',
                '',
                '2. HOW TO MEASURE:',
                '   - Use consistent measurement points across all models',
                '   - Measure diameter at the widest stable point',
                '   - For angle: measure from centerline to centerline',
                '',
                '3. DATA ENTRY:',
                '   - Enter measurements directly in the columns',
                '   - Leave blank if measurement cannot be obtained',
                '   - Use notes column for any special observations',
                '',
                '4. QUALITY CHECKS:',
                '   - Verify measurements are in correct units (mm for diameters, degrees for angle)',
                '   - Check for outliers that might indicate measurement errors',
                '   - Ensure consistency in measurement methodology',
                '',
                '5. WHEN COMPLETE:',
                '   - Save the Excel file',
                '   - Run: python excel_integration_guide.py --process your_filled_file.xlsx',
            ]
        })
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False, header=False)
        
        # Format instructions sheet
        inst_worksheet = writer.sheets['Instructions']
        inst_worksheet.column_dimensions['A'].width = 100
    
    print(f"✅ Created Excel template: {output_file}")
    print("\n📊 Template includes:")
    print(f"   - {len(df)} models to measure")
    print(f"   - Model IDs for easy reference")
    print(f"   - Columns for all three measurements")
    print(f"   - Instructions sheet with guidelines")
    print(f"   - Notes column for observations")
    
    return output_file

def process_excel_measurements(excel_file, sheet_name='Measurements'):
    """
    Process your filled Excel file and integrate with the dataset
    """
    print(f"\n📖 Reading Excel file: {excel_file}")
    
    # Read the Excel file
    df_excel = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # Check which measurements are filled
    measurement_cols = ['left_subclavian_diameter_mm', 'aortic_arch_diameter_mm', 'angle_degrees']
    
    print("\n📊 Data Summary:")
    print(f"Total rows: {len(df_excel)}")
    
    for col in measurement_cols:
        if col in df_excel.columns:
            filled = df_excel[col].notna().sum()
            missing = df_excel[col].isna().sum()
            print(f"\n{col}:")
            print(f"  - Filled: {filled}/{len(df_excel)} ({filled/len(df_excel)*100:.1f}%)")
            print(f"  - Missing: {missing}")
            if filled > 0:
                print(f"  - Range: {df_excel[col].min():.2f} - {df_excel[col].max():.2f}")
                print(f"  - Mean: {df_excel[col].mean():.2f}")
                print(f"  - Std: {df_excel[col].std():.2f}")
    
    # Prepare final dataset
    df_final = df_excel[['filename', 'label']].copy()
    
    # Rename columns to match expected format
    df_final['left_subclavian_diameter'] = df_excel['left_subclavian_diameter_mm']
    df_final['aortic_arch_diameter'] = df_excel['aortic_arch_diameter_mm']
    df_final['angle'] = df_excel['angle_degrees']
    
    # Save processed data
    output_csv = 'classification_labels_with_measurements.csv'
    df_final.to_csv(output_csv, index=False)
    print(f"\n✅ Saved processed data to: {output_csv}")
    
    # Check for data quality issues
    print("\n🔍 Data Quality Checks:")
    
    # Check for outliers
    for col in ['left_subclavian_diameter', 'aortic_arch_diameter', 'angle']:
        if col in df_final.columns and df_final[col].notna().any():
            Q1 = df_final[col].quantile(0.25)
            Q3 = df_final[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = df_final[(df_final[col] < Q1 - 1.5*IQR) | (df_final[col] > Q3 + 1.5*IQR)]
            if len(outliers) > 0:
                print(f"\n⚠️  Potential outliers in {col}:")
                for idx, row in outliers.iterrows():
                    print(f"   {row['filename']}: {row[col]:.2f}")
    
    # Check correlation with labels
    print("\n📈 Correlation with Labels:")
    for col in ['left_subclavian_diameter', 'aortic_arch_diameter', 'angle']:
        if col in df_final.columns and df_final[col].notna().sum() > 0:
            # Only calculate correlation for non-null values
            df_corr = df_final[['label', col]].dropna()
            if len(df_corr) > 1:
                correlation = df_corr.corr().iloc[0, 1]
                print(f"  {col}: {correlation:.3f}")
    
    # Create complete dataset (no missing values)
    df_complete = df_final.dropna(subset=['left_subclavian_diameter', 'aortic_arch_diameter', 'angle'])
    if len(df_complete) < len(df_final):
        complete_csv = 'classification_labels_with_measurements_complete.csv'
        df_complete.to_csv(complete_csv, index=False)
        print(f"\n✅ Created complete dataset (no missing): {complete_csv}")
        print(f"   {len(df_complete)}/{len(df_final)} samples with all measurements")
    
    return df_final

def validate_measurement_ranges(df):
    """
    Validate that measurements are within expected clinical ranges
    """
    print("\n🏥 Clinical Range Validation:")
    
    # Define expected ranges (adjust based on your clinical knowledge)
    ranges = {
        'left_subclavian_diameter': (3, 20),    # mm
        'aortic_arch_diameter': (15, 50),       # mm
        'angle': (20, 180)                      # degrees
    }
    
    warnings = []
    for col, (min_val, max_val) in ranges.items():
        if col in df.columns and df[col].notna().any():
            out_of_range = df[(df[col] < min_val) | (df[col] > max_val)]
            if len(out_of_range) > 0:
                warnings.append(f"{col}: {len(out_of_range)} values outside expected range [{min_val}, {max_val}]")
                print(f"\n⚠️  {col}: {len(out_of_range)} values outside clinical range [{min_val}, {max_val}]")
                for idx, row in out_of_range.head(5).iterrows():
                    print(f"   {row['filename']}: {row[col]:.2f}")
    
    if not warnings:
        print("✅ All measurements within expected clinical ranges")
    
    return warnings

def create_measurement_comparison():
    """
    Create a comparison showing the difference between classes
    """
    if os.path.exists('classification_labels_with_measurements.csv'):
        df = pd.read_csv('classification_labels_with_measurements.csv')
        
        print("\n📊 Measurement Comparison by Class:")
        print("-" * 50)
        
        for label in [0, 1]:
            df_class = df[df['label'] == label]
            print(f"\nClass {label} (n={len(df_class)}):")
            
            for col in ['left_subclavian_diameter', 'aortic_arch_diameter', 'angle']:
                if col in df.columns and df_class[col].notna().any():
                    mean = df_class[col].mean()
                    std = df_class[col].std()
                    print(f"  {col}: {mean:.2f} ± {std:.2f}")
        
        # Statistical tests (if scipy available)
        try:
            from scipy import stats
            print("\n📈 Statistical Significance (t-test):")
            for col in ['left_subclavian_diameter', 'aortic_arch_diameter', 'angle']:
                if col in df.columns:
                    class_0 = df[df['label'] == 0][col].dropna()
                    class_1 = df[df['label'] == 1][col].dropna()
                    if len(class_0) > 0 and len(class_1) > 0:
                        t_stat, p_value = stats.ttest_ind(class_0, class_1)
                        significance = "***" if p_value < 0.001 else "**" if p_value < 0.01 else "*" if p_value < 0.05 else "ns"
                        print(f"  {col}: p={p_value:.4f} {significance}")
        except ImportError:
            print("\n(Install scipy for statistical tests: pip install scipy)")

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Excel Integration Helper for Measurements")
    parser.add_argument('--create', action='store_true', help='Create Excel template')
    parser.add_argument('--process', type=str, help='Process filled Excel file')
    parser.add_argument('--validate', action='store_true', help='Validate existing measurements')
    parser.add_argument('--compare', action='store_true', help='Compare measurements by class')
    
    args = parser.parse_args()
    
    if args.create:
        prepare_excel_template()
        print("\n📝 Next Steps:")
        print("1. Open 'measurement_template.xlsx' in Excel")
        print("2. Fill in the measurement columns for each model")
        print("3. Save the file")
        print("4. Run: python excel_integration_guide.py --process your_file.xlsx")
        
    elif args.process:
        df = process_excel_measurements(args.process)
        validate_measurement_ranges(df)
        create_measurement_comparison()
        print("\n✅ Ready to train!")
        print("Run: python train_with_measurements.py --compare")
        
    elif args.validate:
        if os.path.exists('classification_labels_with_measurements.csv'):
            df = pd.read_csv('classification_labels_with_measurements.csv')
            validate_measurement_ranges(df)
            create_measurement_comparison()
        else:
            print("❌ No measurement file found. Create one first with --create")
            
    elif args.compare:
        create_measurement_comparison()
        
    else:
        # Default: create template
        prepare_excel_template()
        print("\n💡 Tip: Use --help to see all options")